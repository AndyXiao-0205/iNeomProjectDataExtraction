import pandas as pd
import numpy as np
import os
import re
from datetime import datetime
from openpyxl import load_workbook


def try_parse_datetime(date_time_str):
    date_formats = [
        '%b. %d. %Y %H:%M:%S',
        '%m/%d/%Y %H:%M:%S',
        '%Y/%m/%d %H:%M:%S',
        '%b.%d.%Y %H:%M:%S',
        '%d-%b-%Y %H:%M:%S'
    ]
    for date_format in date_formats:
        try:
            return pd.to_datetime(date_time_str, format=date_format)
        except ValueError:
            continue
    return pd.NaT

def process_csv(file_path, output_path):
    try:
        print(f"Processing file: {file_path}")
        data = pd.read_csv(file_path, sep=None, engine='python', on_bad_lines='skip')
        data['Datetime'] = data.apply(lambda row: try_parse_datetime(f"{row.iloc[1]} {row.iloc[2]}"), axis=1)
        data.iloc[:, 3] = pd.to_numeric(data.iloc[:, 3], errors='coerce')
        data = data[data.iloc[:, 3] >= 40]  # Filter out invalid SpO2 data

        if data.empty:
            raise ValueError("No valid SpO2 data found.")

        data['Date'] = data['Datetime'].dt.date
        stats = {
            'Number of Bed': os.path.splitext(os.path.basename(file_path))[0],
            'Number of Days Recorded': data['Date'].nunique(),
            'Number of Minutes Recorded': data['Datetime'].diff().dt.total_seconds().div(60).sum(),
            'Max Value': data.iloc[:, 3].max(),
            'Min Value': data.iloc[:, 3].min(),
            'Average': data.iloc[:, 3].mean(),
            'Median': data.iloc[:, 3].median(),
            'IQR': np.subtract(*np.percentile(data.iloc[:, 3].dropna(), [75, 25])),
            'Total Data Count': len(data)
        }

        df = pd.DataFrame([stats])
    except Exception as e:
        print(f"Error processing {file_path}: {e}")
        df = pd.DataFrame([{
            'Number of Bed': os.path.splitext(os.path.basename(file_path))[0],
            'Number of Days Recorded': "请重新计算本行数据"
        }])

    with pd.ExcelWriter(output_path, mode='a', engine='openpyxl', if_sheet_exists='overlay',) as writer:
        df.to_excel(writer, index=False, header=False,sheet_name='Data', startrow=writer.sheets['Data'].max_row if 'Data' in writer.book.sheetnames else 0)

def process_folder(folder_path, output_path, start_from=104):
    pattern = re.compile(r'cleaned_cleaned_prem_(\d+)\.csv')
    files = sorted(
        (f for f in os.listdir(folder_path) if pattern.match(f)),
        key=lambda x: int(pattern.search(x).group(1))
    )
    for file_name in files:
        file_number = int(pattern.search(file_name).group(1))
        if file_number >= start_from:
            file_path = os.path.join(folder_path, file_name)
            process_csv(file_path, output_path)
    print("All data processed and written to Excel.")

# Define folder path and output path
folder_path = 'D:\\Desktop\\BUPT\\Final Project\\Otras descargas Datos\\spo2\\fileAfterClassify'  # Replace with the path to your folder
output_path = 'D:\\Desktop\\BUPT\\Final Project\\Otras descargas Datos\\spo2\\fileAfterClassify\\outcome_94.xlsx'  # Output Excel file
process_folder(folder_path, output_path)
